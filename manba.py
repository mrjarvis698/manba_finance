import os
from os import path
import shutil
import json
import tkinter
from tkinter import filedialog
import pandas as pd
import json
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException, TimeoutException, NoSuchWindowException
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
import time

root = tkinter.Tk()
root.withdraw()

# Open xlsx file
open_sheet = path.exists("cache/opened_sheet.json")
if open_sheet == True :
  opened_sheet_file_path = "cache/opened_sheet.json"
  json_file = open(opened_sheet_file_path)
  data = json.load(json_file)
  xlsx_sheet_check = path.exists(data ['xlsx_file_path'])
  if xlsx_sheet_check == True :
    xlsx_file_path = data ['xlsx_file_path']
  else :
    shutil.rmtree('cache', ignore_errors=True)
    xlsx_file_path = filedialog.askopenfilename(title="Open Excel-XLSX File")
    cache_path = os.path.join(str(os.getcwd()), "cache")
    dictionary = {"xlsx_file_path" : xlsx_file_path}
    json_object = json.dumps(dictionary, indent = 1)
    with open("cache/opened_sheet.json", "w") as outfile:
      outfile.write(json_object)
else :
  xlsx_file_path = filedialog.askopenfilename(title="Open Excel-XLSX File")
  cache_path = os.path.join(str(os.getcwd()), "cache")
  os.mkdir(cache_path)
  dictionary = {"xlsx_file_path" : xlsx_file_path}
  json_object = json.dumps(dictionary, indent = 1)
  with open("cache/opened_sheet.json", "w") as outfile:
    outfile.write(json_object)

# read imported xlsx file path using pandas
input_workbook = pd.read_excel(xlsx_file_path, sheet_name = 'Sheet1', usecols = 'E:I', dtype=str)
input_workbook.head()

# read total number of rows present in xlsx
number_of_rows = len(input_workbook.index)

# Opening JSON file & returns JSON object as a dictionary
json_file = open('settings.json')
settings_data = json.load(json_file)

input_workbook_cc_number = input_workbook['Card Number'].values.tolist()
input_workbook_cvv_number = input_workbook['CVV'].values.tolist()
input_workbook_expiry_number = input_workbook['Expiry'].values.tolist()
input_workbook_ipin  = input_workbook['Ipin'].values.tolist()
input_workbook_desk_number = input_workbook['Desk'].values.tolist()

# get-output sheet to append output
output_sheet = path.exists("Output.xlsx")
if output_sheet == True :
  output_sheet_file_path = "Output.xlsx"
else :
  output_headers= ['FirstName','LastName', 'Mobile', 'Email','Amount', 'CardNumber', 'Ipin', 'CVV', 'Expiry', 'Status', 'Transation No', 'No.of Transactions', 'Desk', "Desk Holder"]
  overall_output = Workbook()
  page = overall_output.active
  page.append(output_headers)
  overall_output.save(filename = 'Output.xlsx')
  output_sheet_file_path = "Output.xlsx"

def cal():
  global output_cc_number
  global done_transactions_wb_1
  global h
  output_load_wb_2 = pd.read_excel(output_sheet_file_path, sheet_name = 'Sheet', usecols = 'F', dtype=str)
  output_load_wb_2.head()
  output_cc_number = output_load_wb_2['CardNumber'].values.tolist()
  output_load_wb_1 = pd.read_excel(output_sheet_file_path, sheet_name = 'Sheet', usecols = 'L', dtype=int)
  output_load_wb_1.head()
  done_transactions_wb_1 = output_load_wb_1['No.of Transactions'].values.tolist()
  h = len(output_load_wb_1.index) - 1
  print (output_cc_number[h],done_transactions_wb_1[h])

def cc_expiry():
  global expiry_month
  global expiry_year
  global expiry_year1
  global expiry_year2
  global expiry_year3
  global expiry_year4
  workbook_expiry_month = input_workbook_expiry_number[x]
  workbook_expiry_year = input_workbook_expiry_number[x]
  expiry_month = workbook_expiry_month[:2]
  expiry_year = workbook_expiry_year[5:]
  expiry_year1 = workbook_expiry_year[3]
  expiry_year2 = workbook_expiry_year[4]
  expiry_year3 = workbook_expiry_year[5]
  expiry_year4 = workbook_expiry_year[6]
def start_link():
  driver.get("https://mfq.manbafinance.com/paymentwebsite")

def page_one():
  # LAN Text Box
  try :
    WebDriverWait(driver, timeout=8).until(ec.visibility_of_element_located((By.XPATH, '//*[@id="txtlanno"]')))
  except TimeoutException:
    timeout_exception = True
    timeout_exception1 = True
  else :
    timeout_exception = False
    timeout_exception1 = False
    textbox_elements = driver.find_element_by_xpath ('//*[@id="txtlanno"]')
    textbox_elements.send_keys(settings_data['LAN'])

  # LAN Next Page
  try :
    WebDriverWait(driver, timeout=8).until(ec.visibility_of_element_located((By.XPATH, '//*[@id="next"]')))
  except TimeoutException:
    timeout_exception = True
    timeout_exception1 = True
  else :
    timeout_exception = False
    page_button = driver.find_element_by_xpath ('//*[@id="next"]')
    page_button.click()

def page_two():
  # Select EMI Amount
  try :
    WebDriverWait(driver, timeout=8).until(ec.visibility_of_element_located((By.XPATH, '//*[@id="Other"]')))
  except TimeoutException:
    timeout_exception = True
    timeout_exception1 = True
  else :
    timeout_exception = False
    timeout_exception1 = False
    radio_button = driver.find_element_by_xpath ('//*[@id="Other"]')
    radio_button.click()

  # Overdue Interest
  try :
    WebDriverWait(driver, timeout=8).until(ec.visibility_of_element_located((By.XPATH, '//*[@id="txtamount"]')))
  except TimeoutException:
    timeout_exception = True
    timeout_exception1 = True
  else :
    timeout_exception = False
    page_button = driver.find_element_by_xpath ('//*[@id="txtamount"]')
    page_button.send_keys(settings_data['payable_amount'])
  
  # Make Payment Button
  try :
    WebDriverWait(driver, timeout=8).until(ec.visibility_of_element_located((By.XPATH, '//*[@id="btnPay"]')))
  except TimeoutException:
    timeout_exception = True
    timeout_exception1 = True
  else :
    timeout_exception = False
    timeout_exception1 = False
    submit_button = driver.find_element_by_xpath ('//*[@id="btnPay"]')
    submit_button.click()

def page_extra():
  driver.switch_to.frame(WebDriverWait(driver, timeout=10).until(ec.visibility_of_element_located((By.CLASS_NAME, "paymtiframe"))))

  # Proceed Button
  try :
    WebDriverWait(driver, timeout=8).until(ec.visibility_of_element_located((By.XPATH, '//*[@id="mob-payment-btn"]')))
  except TimeoutException:
    timeout_exception = True
    timeout_exception1 = True
  else :
    timeout_exception = False
    timeout_exception1 = False
    submit_button = driver.find_element_by_xpath ('//*[@id="mob-payment-btn"]')
    submit_button.click()

def page_three():
  
  driver.switch_to.frame(WebDriverWait(driver, timeout=10).until(ec.visibility_of_element_located((By.CLASS_NAME, "razorpay-checkout-frame"))))
  # Phone
  try :
    WebDriverWait(driver, timeout=10).until(ec.visibility_of_element_located((By.XPATH, '//*[@id="contact"]')))
  except TimeoutException:
    timeout_exception = True
    timeout_exception1 = True
  else :
    timeout_exception = False
    page_button = driver.find_element_by_xpath ('//*[@id="contact"]')
    page_button.send_keys(settings_data['registered_mobile_no'])

  # Email
  try :
    WebDriverWait(driver, timeout=10).until(ec.visibility_of_element_located((By.XPATH, '//*[@id="email"]')))
  except TimeoutException:
    timeout_exception = True
    timeout_exception1 = True
  else :
    timeout_exception = False
    page_button = driver.find_element_by_xpath ('//*[@id="email"]')
    page_button.send_keys(settings_data['email_id'])
  
  # Proceed Button
  try :
    WebDriverWait(driver, timeout=10).until(ec.visibility_of_element_located((By.XPATH, '//*[@id="footer-cta"]')))
  except TimeoutException:
    timeout_exception = True
    timeout_exception1 = True
  else :
    timeout_exception = False
    timeout_exception1 = False
    submit_button = driver.find_element_by_xpath ('//*[@id="footer-cta"]')
    submit_button.click()

def page_four():
  # Card Type
  try:
      WebDriverWait(driver, timeout=8).until(ec.visibility_of_element_located((By.XPATH, '//*[@id="form-common"]/div[1]/div/div/div[2]/div[1]/div/button[1]/div/div[1]/div[1]')))
  except TimeoutException:
    timeout_exception = True
    timeout_exception1 = True
  else :
    card_type1 = WebDriverWait(driver, timeout=8).until(ec.visibility_of_element_located((By.XPATH, '//*[@id="form-common"]/div[1]/div/div/div[2]/div[1]/div/button[1]/div/div[1]/div[1]')))
    timeout_exception = False
    timeout_exception1 = False
  
  try:
      WebDriverWait(driver, timeout=8).until(ec.visibility_of_element_located((By.XPATH, '//*[@id="form-common"]/div[1]/div/div/div[2]/div[2]/div/button[1]/div/div[1]/div[1]')))
  except TimeoutException:
    timeout_exception = True
    timeout_exception1 = True
  else :
    card_type2 = WebDriverWait(driver, timeout=8).until(ec.visibility_of_element_located((By.XPATH, '//*[@id="form-common"]/div[1]/div/div/div[2]/div[2]/div/button[1]/div/div[1]/div[1]')))
    timeout_exception = False
    timeout_exception1 = False
  
  if card_type1.text == 'Pay using Card':
    card_type1.click()
  else :
    card_type2.click()

def page_five():
  # CC number
  try :
    WebDriverWait(driver, timeout=10).until(ec.visibility_of_element_located((By.XPATH, '//*[@id="card_number"]')))
  except TimeoutException:
    timeout_exception = True
    timeout_exception1 = True
  else :
    timeout_exception = False
    page_button = driver.find_element_by_xpath ('//*[@id="card_number"]')
    page_button.send_keys(input_workbook_cc_number[x])
  
  # Expiry
  cc_expiry()
  try :
    WebDriverWait(driver, timeout=10).until(ec.visibility_of_element_located((By.XPATH, '//*[@id="card_expiry"]')))
  except TimeoutException:
    timeout_exception = True
    timeout_exception1 = True
  else :
    timeout_exception = False
    page_button = driver.find_element_by_xpath ('//*[@id="card_expiry"]')
    page_button.send_keys(expiry_month + expiry_year3 + expiry_year4)
  
  # CC holder name
  try :
    WebDriverWait(driver, timeout=10).until(ec.visibility_of_element_located((By.XPATH, '//*[@id="card_name"]')))
  except TimeoutException:
    timeout_exception = True
    timeout_exception1 = True
  else :
    timeout_exception = False
    page_button = driver.find_element_by_xpath ('//*[@id="card_name"]')
    page_button.send_keys(settings_data['first_name'])
  
  # CVV
  try :
    WebDriverWait(driver, timeout=10).until(ec.visibility_of_element_located((By.XPATH, '//*[@id="card_cvv"]')))
  except TimeoutException:
    timeout_exception = True
    timeout_exception1 = True
  else :
    timeout_exception = False
    page_button = driver.find_element_by_xpath ('//*[@id="card_cvv"]')
    page_button.send_keys(input_workbook_cvv_number[x])
  
  # Proceed
  try :
    WebDriverWait(driver, timeout=10).until(ec.visibility_of_element_located((By.XPATH, '//*[@id="footer-cta"]')))
  except TimeoutException:
    timeout_exception = True
    timeout_exception1 = True
  else :
    timeout_exception = False
    timeout_exception1 = False
    submit_button = driver.find_element_by_xpath ('//*[@id="footer-cta"]')
    submit_button.click()

def page_six():
  driver.switch_to.window(driver.window_handles[1])
  # IPIN
  try :
    WebDriverWait(driver, timeout=10).until(ec.visibility_of_element_located((By.XPATH, '//*[@id="ipin"]')))
  except TimeoutException:
    timeout_exception = True
    timeout_exception1 = True
  else :
    timeout_exception = False
    page_button = driver.find_element_by_xpath ('//*[@id="ipin"]')
    page_button.send_keys(input_workbook_ipin[x])
  
  # Ipin Submit
  try :
    WebDriverWait(driver, timeout=10).until(ec.visibility_of_element_located((By.XPATH, '//*[@id="otpbut"]')))
  except TimeoutException:
    timeout_exception = True
    timeout_exception1 = True
  else :
    timeout_exception = False
    timeout_exception1 = False
    submit_button = driver.find_element_by_xpath ('//*[@id="otpbut"]')
    submit_button.click()

def output():
  global output_status
  global transaction_output_status
  time.sleep(1)
  try :
    WebDriverWait(driver, timeout=4).until(ec.visibility_of_element_located((By.XPATH, '//*[@id="set"]/div/div/div[2]/div/div[3]/font')))
  except (TimeoutException, NoSuchWindowException):
    driver.switch_to.window(driver.window_handles[0])
    driver.switch_to.frame(WebDriverWait(driver, timeout=8).until(ec.visibility_of_element_located((By.CLASS_NAME, "paymtiframe"))))

    #driver.switch_to.frame(WebDriverWait(driver, timeout=8).until(ec.visibility_of_element_located((By.CLASS_NAME, "razorpay-checkout-frame"))))
    try :
      WebDriverWait(driver, timeout=3).until(ec.visibility_of_element_located((By.XPATH, '//*[@id="checkout-parent"]/div[2]/div[2]/div')))
    except TimeoutException :
      try:
        WebDriverWait(driver, timeout=8).until(ec.visibility_of_element_located((By.XPATH, '//*[@id="fd-t"]')))
      except TimeoutException:
        timeout_exception = True
        timeout_exception1 = True
      else:
        output_status = "Failed"
        transaction_output_status = '-'
        timeout_exception = False
        timeout_exception1 = False
    else :
      output_status = "Success"
      transaction_element = driver.find_element_by_xpath('//*[@id="checkout-parent"]/div[2]/div[2]/div')
      transaction_output_status = transaction_element.text
      timeout_exception = False
      timeout_exception1 = False
  else :
    try :
      driver.find_element_by_xpath ('//*[@id="cancel"]')
    except NoSuchElementException:
      driver.find_element_by_xpath ('//*[@id="cancel"]').click()
    else :
      driver.find_element_by_xpath ('//*[@id="cancel"]').click()
    driver.switch_to.window(driver.window_handles[0])
    driver.switch_to.frame(WebDriverWait(driver, timeout=8).until(ec.visibility_of_element_located((By.CLASS_NAME, "paymtiframe"))))
    driver.switch_to.frame(WebDriverWait(driver, timeout=8).until(ec.visibility_of_element_located((By.CLASS_NAME, "razorpay-checkout-frame"))))
    output_status = "Please enter correct IPIN / WEB PIN"
    transaction_output_status = "-"

def output_save():
  entry_list = [[settings_data['first_name'], settings_data['last_name'], settings_data['registered_mobile_no'], settings_data['email_id'], settings_data['payable_amount'], input_workbook_cc_number[x], input_workbook_ipin[x], input_workbook_cvv_number[x], input_workbook_expiry_number[x], output_status, transaction_output_status, z+1, int(input_workbook_desk_number[x]), settings_data["desk_holder"]]]
  output_wb = load_workbook(output_sheet_file_path)
  page = output_wb.active
  for info in entry_list:
      page.append(info)
  output_wb.save(filename='Output.xlsx')

def whole_work():
    start_link()
    page_one()
    page_two()
    page_extra()
    page_three()
    page_four()
    page_five()
    page_six()
    output()
    output_save()

chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument("--incognito")
caps = DesiredCapabilities().CHROME
#caps["pageLoadStrategy"] = "none"
#caps["pageLoadStrategy"] = "eager"
caps["pageLoadStrategy"] = "normal"
driver=webdriver.Chrome(chrome_options=chrome_options, desired_capabilities=caps, executable_path="chromedriver.exe")
#driver.maximize_window()
try:
  cal()
except IndexError:
  for x in range (0 , number_of_rows):
    for z in range (0, int(settings_data['number_of_time_transactions_per_card'])):
      whole_work()
else:
  last_txncard =  input_workbook[input_workbook['Card Number'] == output_cc_number[h]].index[0]
  for x in range (last_txncard , number_of_rows):
    for z in range (done_transactions_wb_1[h], int(settings_data['number_of_time_transactions_per_card'])):
      whole_work()
    done_transactions_wb_1[h] = 0

driver.quit()
